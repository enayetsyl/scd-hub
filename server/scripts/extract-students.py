#!/usr/bin/env python3
"""
Extract the real student roster from the source .xlsx into a normalized JSON the
TypeScript loader (import-students.ts) consumes. Identity-plane PII — both the
source .xlsx and the emitted JSON are gitignored (ADR-005, D-#31).

Usage (from repo root):
    python server/scripts/extract-students.py \
        "Students - School for Community Development.xlsx" \
        server/scripts/students.json

Mapping decisions (D-#30/#31):
  - Class name -> ROSTER class level: Nursery=-1, KG=0, One..Five=1..5.
  - Section "Boys"/"Girls" kept as-is; blank -> default "Main".
  - Primary phone = "SMS Contact"; parents captured as login-disabled guardian contacts.
  - Phone strings carry zero-width chars in the source; they are stripped.
"""
import json
import re
import sys

import openpyxl

CLASS_LEVELS = {
    "nursery": -1,
    "kg": 0,
    "one": 1,
    "two": 2,
    "three": 3,
    "four": 4,
    "five": 5,
}

GENDERS = {"male": "male", "female": "female"}


def clean(v):
    """Trim, drop zero-width chars, collapse whitespace; '' -> None."""
    if v is None:
        return None
    s = str(v)
    s = s.replace("​", "").replace("‌", "").replace("‍", "").replace("﻿", "")
    s = s.strip()
    return s or None


def clean_phone(v):
    s = clean(v)
    if not s:
        return None
    # keep a leading +, drop spaces/dashes/parens inside
    s = re.sub(r"(?!^\+)[^\d]", "", s)
    return s or None


def to_iso_date(v):
    """Source DOB is either a datetime or a 'DD-MM-YYYY' string."""
    if v is None:
        return None
    # openpyxl may already give a datetime
    if hasattr(v, "isoformat") and hasattr(v, "year"):
        return v.date().isoformat() if hasattr(v, "date") else v.isoformat()
    s = clean(v)
    if not s:
        return None
    m = re.match(r"^(\d{1,2})-(\d{1,2})-(\d{4})$", s)
    if m:
        d, mo, y = m.groups()
        return f"{y}-{int(mo):02d}-{int(d):02d}"
    return None  # unparseable -> drop rather than guess


def main():
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(2)
    src, dst = sys.argv[1], sys.argv[2]

    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb["Sheet1"]
    header = [clean(c.value) for c in ws[2]]
    idx = {h: i for i, h in enumerate(header) if h}

    def cell(row, name):
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else None

    students = []
    skipped = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        school_id = clean(cell(row, "ID"))
        name = clean(cell(row, "Name"))
        class_name = clean(cell(row, "Class"))
        if not school_id or not name or not class_name:
            continue  # blank/padding row
        level = CLASS_LEVELS.get(class_name.lower())
        if level is None:
            skipped.append((school_id, name, class_name))
            continue

        gender_raw = clean(cell(row, "Gender"))
        gender = GENDERS.get(gender_raw.lower()) if gender_raw else None

        guardians = []
        for relation, name_col, phone_col in [
            ("father", "Father Name", "Father Contact"),
            ("mother", "Mother Name", "Mother Contact"),
            ("guardian", "Guardian Name", "Guardian Contact"),
        ]:
            gname = clean(cell(row, name_col))
            gphone = clean_phone(cell(row, phone_col))
            if gname or gphone:
                guardians.append({"relation": relation, "name": gname, "phone": gphone})

        students.append({
            "schoolId": school_id,
            "name": name,
            "nameBn": clean(cell(row, "Name (Bangla)")),
            "classLevel": level,
            "section": clean(cell(row, "Section")),  # None | "Boys" | "Girls"
            "gender": gender,
            "dob": to_iso_date(cell(row, "Date of Birth")),
            "phone": clean_phone(cell(row, "SMS Contact")),
            "address": clean(cell(row, "Present Address")) or clean(cell(row, "Permanent Address")),
            "bloodGroup": clean(cell(row, "Blood Group")),
            "guardians": guardians,
        })

    out = {"students": students}
    with open(dst, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    print(f"Extracted {len(students)} students -> {dst}")
    if skipped:
        print(f"  WARNING: {len(skipped)} rows had an unrecognized Class and were skipped:")
        for sid, nm, cn in skipped:
            print(f"    {sid}  {nm}  class={cn!r}")


if __name__ == "__main__":
    main()
