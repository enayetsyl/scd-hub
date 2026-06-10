#!/usr/bin/env python3
"""
Extract the real staff roster (teachers + administrative staff) from the source
.xlsx files into a normalized JSON the TypeScript loader (import-staff.ts)
consumes. Identity-plane PII — both the source .xlsx and the emitted JSON are
gitignored (ADR-005; prd-hr H1.4 — sensitive rows like NID/bank are gated).

Usage (from repo root):
    python server/scripts/extract-staff.py [out.json]
    # default out: server/scripts/staff.json

The two source workbooks live in the repo root; each maps to one HR category
(prd-hr §2.2). Edit SOURCES if the filenames change.

Mapping decisions (HR-1):
  - Contract Type -> employmentType: Full-Time=full_time, Part-Time=part_time.
  - employmentStatus defaults to "confirmed" (no Termination Date in source).
  - Phone strings carry zero-width chars + a leading +; they are normalized.
  - Attendance Device ID -> biometricId (the future attendance mapping key, H1.5).
  - phone = Contact (SMS), else Contact (Personal).
"""
import json
import os
import re
import sys

import openpyxl

# (filename in repo root, HR category) — the one-off real-roster load.
SOURCES = [
    ("Teachers - School for Community Development (1).xlsx", "teacher"),
    ("Administratives - School for Community Development.xlsx", "office_accounts"),
]

EMPLOYMENT_TYPE = {"full-time": "full_time", "part-time": "part_time", "fixed-term": "fixed_term"}
GENDERS = {"male": "male", "female": "female"}


def clean(v):
    """Trim, drop zero-width chars, collapse whitespace; '' -> None."""
    if v is None:
        return None
    s = str(v)
    s = s.replace("​", "").replace("‌", "").replace("‍", "").replace("﻿", "")
    s = re.sub(r"\s+", " ", s).strip()
    return s or None


def clean_phone(v):
    s = clean(v)
    if not s:
        return None
    s = re.sub(r"(?!^\+)[^\d]", "", s)
    return s or None


def to_iso_date(v):
    """Source date is either a datetime or a 'DD-MM-YYYY' string."""
    if v is None:
        return None
    if hasattr(v, "isoformat") and hasattr(v, "year"):
        return v.date().isoformat() if hasattr(v, "date") else v.isoformat()
    s = clean(v)
    if not s:
        return None
    m = re.match(r"^(\d{1,2})-(\d{1,2})-(\d{4})$", s)
    if m:
        d, mo, y = m.groups()
        return f"{y}-{int(mo):02d}-{int(d):02d}"
    return None  # unparseable -> drop rather than guess


def emp_type(raw):
    s = clean(raw)
    return EMPLOYMENT_TYPE.get(s.lower(), "full_time") if s else "full_time"


def extract_file(path, category):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    header = [clean(c.value) for c in ws[2]]  # row 1 is a banner title; row 2 = headers
    idx = {h: i for i, h in enumerate(header) if h}

    def cell(row, name):
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else None

    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        school_id = clean(cell(row, "ID"))
        name = clean(cell(row, "Name"))
        if not school_id or not name:
            continue  # blank/padding row

        gender_raw = clean(cell(row, "Gender"))
        gender = GENDERS.get(gender_raw.lower()) if gender_raw else None

        rows.append({
            "schoolId": school_id,
            "name": name,
            "nameBn": clean(cell(row, "Name (Bangla)")),
            "category": category,
            "designation": clean(cell(row, "Designation")),
            "employmentType": emp_type(cell(row, "Contract Type")),
            "employmentStatus": "confirmed",
            "joiningDate": to_iso_date(cell(row, "Joining Date")),
            "biometricId": clean(cell(row, "Attendance Device ID")),
            "gender": gender,
            "dob": to_iso_date(cell(row, "Date of Birth")),
            "bloodGroup": clean(cell(row, "Blood Group")),
            "maritalStatus": clean(cell(row, "Martial Status")),
            "nationality": clean(cell(row, "Nationality")),
            "qualification": clean(cell(row, "Qualification")),
            "majoredIn": clean(cell(row, "Majored in")),
            "studiedAt": clean(cell(row, "Studied at")),
            "fatherName": clean(cell(row, "Father's Name")),
            "motherName": clean(cell(row, "Mother's Name")),
            "spouseName": clean(cell(row, "Spouse's Name")),
            "phone": clean_phone(cell(row, "Contact (SMS)")) or clean_phone(cell(row, "Contact (Personal)")),
            "whatsapp": clean_phone(cell(row, "Whatsapp")),
            "email": clean(cell(row, "Email")),
            "presentAddress": clean(cell(row, "Present Address")),
            "permanentAddress": clean(cell(row, "Permanent Address")),
            "nid": clean(cell(row, "NID No.")),
            "bankAccount": clean(cell(row, "Bank Account")),
        })
    return rows


def main():
    dst = sys.argv[1] if len(sys.argv) > 1 else "server/scripts/staff.json"
    staff = []
    for fname, category in SOURCES:
        if not os.path.exists(fname):
            print(f"  WARNING: source not found, skipped: {fname!r}")
            continue
        rows = extract_file(fname, category)
        print(f"  {len(rows):>3} {category:<16} <- {fname}")
        staff.extend(rows)

    # guard against duplicate schoolIds across files
    seen = {}
    for s in staff:
        seen.setdefault(s["schoolId"], []).append(s["name"])
    dupes = {k: v for k, v in seen.items() if len(v) > 1}

    with open(dst, "w", encoding="utf-8") as f:
        json.dump({"staff": staff}, f, ensure_ascii=False, indent=2)

    print(f"Extracted {len(staff)} staff -> {dst}")
    if dupes:
        print(f"  WARNING: {len(dupes)} duplicate schoolId(s): {dupes}")


if __name__ == "__main__":
    main()
